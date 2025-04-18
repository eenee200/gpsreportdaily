import os
import json
import pandas as pd
from datetime import datetime, timedelta
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.application import MIMEApplication
from email.mime.text import MIMEText
from openpyxl import Workbook
from openpyxl.styles import Border, Side, Alignment, Font, PatternFill
from openpyxl.formatting.rule import FormulaRule
import requests
RECEIVER_EMAILS="uuganbileg@tttools.mn,lhagvabayar.a@monospharmatrade.mn,undrakh.b@monospharmatrade.mn,anuerdene.b@monospharmatrade.mn,ayurzana.s@monospharmatrade.mn,munkhtamir.b@monospharmatrade.mn,baatarkhuu@monospharmatrade.mn,ariuntungalag@monospharmatrade.mn,narmandakh.b@monospharmatrade.mn"

# Configuration (Replace with your actual details)
CONFIG = {
    'GPS_API_KEY': os.environ.get('GPS_API_KEY'),
    'VEHICLES': {
       '866069068751245': '2514УАС ОН',
        '867747074781160': '3049УАС МУБ',
        '866069068899358': '6461УНЯ ОН',
        '866069068945011': '7107УБГ ОН',
        '866069064383431': '7228УКР МУБ',
        '867747074781186': '7228УКС МУБ'
    },
    'SENDER_EMAIL': os.environ.get('SENDER_EMAIL'),
    'SENDER_PASSWORD': os.environ.get('SENDER_PASSWORD'),
    'RECEIVER_EMAILS': RECEIVER_EMAILS.split(','),
    'REPORT_FREQUENCY_DAYS': 1
}
def get_nearest_temperatures(timestamp, temp_data, target_hours=[10, 12, 15], time_threshold=60):
    """
    Find nearest temperature readings for specified hours relative to a given timestamp.
    If no reading found within threshold, returns the closest previous temperature reading.
    """
    date = timestamp.date()
    result = {}
    
    # Filter temperature data for the same date
    same_date_readings = [
        reading for reading in temp_data 
        if reading['timestamp'].date() == date
    ]
    
    # Sort readings by timestamp
    same_date_readings.sort(key=lambda x: x['timestamp'])
    
    for target_hour in target_hours:
        # Create target timestamp for comparison
        target_time = datetime.combine(date, datetime.min.time().replace(hour=target_hour))
        
        # First try to find reading within threshold
        nearest_temp = next(
            (reading['temperature'] for reading in same_date_readings
            if abs((reading['timestamp'] - target_time).total_seconds()) < time_threshold),
            None
        )
        
        # If no reading found within threshold, find the closest previous reading
        if nearest_temp is None:
            previous_readings = [
                reading for reading in same_date_readings
                if reading['timestamp'] < target_time
            ]
            
            if previous_readings:
                # Get the most recent previous reading
                nearest_temp = previous_readings[-1]['temperature']
            else:
                nearest_temp = 0  # No previous readings found
                
        result[target_hour] = nearest_temp
    
    return result

def parse_gps_temp_door(json_data, plate_number):
    """
    Parse GPS tracking data to extract bag temperature, storage temperature, and door status.
    Now includes specific plate number for the vehicle.
    """
    sensor_data = {
        'storage_temp': [],
        'bag_temp': [],
        'door': [],
        'plate_number': plate_number
    }
    
    last_valid_storage_temp = 0
    last_valid_bag_temp = 0
    
    for entry in json_data:
        timestamp = datetime.strptime(entry[0], '%Y-%m-%d %H:%M:%S') + timedelta(hours=8)
        
        if 'io10800' in entry[6]:
            storage_temp = float(entry[6]['io10800']) / 100
            if storage_temp == 250:
                storage_temp = last_valid_storage_temp
            else:
                last_valid_storage_temp = storage_temp
            sensor_data['storage_temp'].append({
                'timestamp': timestamp,
                'temperature': storage_temp
            })
        
        if 'io10801' in entry[6]:
            bag_temp = float(entry[6]['io10801']) / 100
            if bag_temp == 250:
                bag_temp = last_valid_bag_temp
            else:
                last_valid_bag_temp = bag_temp
            sensor_data['bag_temp'].append({
                'timestamp': timestamp,
                'temperature': bag_temp
            })
        
        if 'io10808' in entry[6]:
            door_status = 1 if float(entry[6]['io10808']) == 250 else 0
            sensor_data['door'].append([
                int(timestamp.timestamp() * 1000),
                door_status
            ])
    
    return sensor_data

def process_temperature_data(temperature_data, target_hours=[10, 12, 15], time_threshold=60):
    """
    Process temperature data for specific target hours with improved nearest reading finding
    """
    daily_temps = {}
    
    # Group readings by date
    date_groups = {}
    for entry in temperature_data:
        date = entry['timestamp'].date()
        if date not in date_groups:
            date_groups[date] = []
        date_groups[date].append(entry)
    
    # Find nearest readings for each target hour
    for date, readings in date_groups.items():
        daily_temps[date] = get_nearest_temperatures(
            datetime.combine(date, datetime.min.time()),
            readings,
            target_hours,
            time_threshold
        )
    
    return daily_temps
def process_door_events(door_data, storage_temp_data, bag_temp_data):
    """
    Process door sensor data to find complete door opening periods
    """
    daily_door_events = {}
    last_state = 0
    first_open_time = None
    
    for timestamp_ms, state in door_data:
        timestamp = datetime.fromtimestamp(timestamp_ms / 1000)
        date_key = timestamp.date()
        
        if state == 1 and last_state == 0:
            first_open_time = timestamp
        elif state == 0 and first_open_time is not None:
            if date_key not in daily_door_events:
                daily_door_events[date_key] = []
            
            # Find nearest temperatures
            activation_storage = next((t['temperature'] for t in storage_temp_data 
                if abs((t['timestamp'] - first_open_time).total_seconds()) < 1), 0)
            deactivation_storage = next((t['temperature'] for t in storage_temp_data 
                if abs((t['timestamp'] - timestamp).total_seconds()) < 1), 0)
            activation_bag = next((t['temperature'] for t in bag_temp_data 
                if abs((t['timestamp'] - first_open_time).total_seconds()) < 1), 0)
            deactivation_bag = next((t['temperature'] for t in bag_temp_data 
                if abs((t['timestamp'] - timestamp).total_seconds()) < 1), 0)
            
            daily_door_events[date_key].append({
                'activation_time': first_open_time,
                'deactivation_time': timestamp,
                'activation_storage_temp': activation_storage,
                'deactivation_storage_temp': deactivation_storage,
                'activation_bag_temp': activation_bag,
                'deactivation_bag_temp': deactivation_bag,
                'duration': timestamp - first_open_time
            })
            first_open_time = None
        
        last_state = state
    
    return daily_door_events

def export_to_excel(vehicles_data, output_file='temperature_analysis.xlsx'):
    """
    Export analyzed data to Excel with support for multiple vehicles, including blank entries for vehicles with no data
    """
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Temperature Analysis"
    
    headers_1 = [
        'Plate Number', 'Date',
        'Storage Temp 10:00', 'Storage Temp 12:00', 'Storage Temp 15:00',
        'Bag Temp 10:00', 'Bag Temp 12:00', 'Bag Temp 15:00'
    ]
    
    headers_2 = [
        'Daily Door Events', 'Total Events',
        'Door Open Time', 'Storage Temp at Open', 'Bag Temp at Open',
        'Door Close Time', 'Storage Temp at Close', 'Bag Temp at Close',
        'Duration'
    ]
    
    sheet.append(headers_1 + headers_2)
    
    # Styling
    border = Border(left=Side(style='thin'), right=Side(style='thin'),
                   top=Side(style='thin'), bottom=Side(style='thin'))
    alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')
    
    start_row = 2
    total_events = 0
    
    # Calculate total events across all vehicles
    for vehicle_data in vehicles_data:
        for date_events in vehicle_data['door_events'].values():
            total_events += len(date_events)
    
    # Get all unique dates across all vehicles
    all_dates = set()
    for vehicle_data in vehicles_data:
        dates = set()
        dates.update(vehicle_data['storage_temps'].keys())
        dates.update(vehicle_data['bag_temps'].keys())
        dates.update(vehicle_data['door_events'].keys())
        all_dates.update(dates)
    
    # If no dates found, add current date
    if not all_dates:
        all_dates.add(datetime.now().date())
    
    # Sort dates
    all_dates = sorted(all_dates)
    
    # Process each vehicle's data
    for vehicle_data in vehicles_data:
        plate_number = vehicle_data['plate_number']
        storage_temps = vehicle_data.get('storage_temps', {})
        bag_temps = vehicle_data.get('bag_temps', {})
        door_events = vehicle_data.get('door_events', {})
        
        # Add at least one row for vehicles with no data
        if not storage_temps and not bag_temps and not door_events:
            row = [
                plate_number,  # Plate number
                all_dates[0].strftime('%Y/%m/%d'),  # Use first date
                '', '', '',  # Empty storage temps
                '', '', '',  # Empty bag temps
                '0', '',  # No door events
                '', '', '',  # Empty door open data
                '', '', '',  # Empty door close data
                ''  # Empty duration
            ]
            sheet.append(row)
            start_row += 1
            continue
        
        is_first_date = True
        for date in all_dates:
            storage_data = storage_temps.get(date, {})
            bag_data = bag_temps.get(date, {})
            daily_events = door_events.get(date, [])
            
            base_row = [
                plate_number if is_first_date else '',
                date.strftime('%Y/%m/%d'),
                storage_data.get(10, ''), storage_data.get(12, ''), storage_data.get(15, ''),
                bag_data.get(10, ''), bag_data.get(12, ''), bag_data.get(15, ''),
            ]
            
            if daily_events:
                first_event = daily_events[0]
                row = base_row + [
                    len(daily_events),
                    total_events if is_first_date else '',
                    first_event['activation_time'].strftime('%H:%M:%S'),
                    first_event['activation_storage_temp'],
                    first_event['activation_bag_temp'],
                    first_event['deactivation_time'].strftime('%H:%M:%S'),
                    first_event['deactivation_storage_temp'],
                    first_event['deactivation_bag_temp'],
                    str(first_event['duration'])
                ]
                sheet.append(row)
                
                for event in daily_events[1:]:
                    row = [''] * 8 + [
                        '',
                        '',
                        event['activation_time'].strftime('%H:%M:%S'),
                        event['activation_storage_temp'],
                        event['activation_bag_temp'],
                        event['deactivation_time'].strftime('%H:%M:%S'),
                        event['deactivation_storage_temp'],
                        event['deactivation_bag_temp'],
                        str(event['duration'])
                    ]
                    sheet.append(row)
                
                if len(daily_events) > 1:
                    end_row = sheet.max_row
                    sheet.row_dimensions.group(start_row + 1, end_row, outline_level=1, hidden=True)
                    start_row = end_row + 1
                else:
                    start_row += 1
            else:
                row = base_row + ['0', '', '', '', '', '', '', '', '']
                sheet.append(row)
                start_row += 1
            
            is_first_date = False
    
    # Apply formatting and styling
    for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=len(headers_1) + len(headers_2)):
        for cell in row:
            cell.border = border
            cell.alignment = alignment
    
    # Adjust column widths
    for col in sheet.columns:
        max_length = 0
        for cell in col:
            try:
                max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        sheet.column_dimensions[col[0].column_letter].width = max_length + 2
    
    # Temperature highlighting
    red_fill = PatternFill(start_color='FFFF0000', end_color='FFFF0000', fill_type='solid')
    temp_columns = ['C', 'D', 'E', 'F', 'G', 'H', 'L', 'M', 'O', 'P']
    
    for col in temp_columns:
        sheet.conditional_formatting.add(
            f'{col}2:{col}{sheet.max_row}',
            FormulaRule(formula=[f'AND(ISNUMBER({col}2), {col}2 <= -9)'],
                       stopIfTrue=True, fill=red_fill)
        )
    
    workbook.save(output_file)
    return output_file
def send_email_with_attachment(sender_email, sender_password, receiver_emails, 
                             subject, message, attachment_path):
    """
    Send email with Excel report attachment to multiple recipients
    
    Parameters:
    -----------
    sender_email : str
        Email address of the sender
    sender_password : str
        Password for the sender's email account
    receiver_emails : list or str
        List of recipient email addresses or a single email address
    subject : str
        Subject of the email
    message : str
        Body of the email
    attachment_path : str
        Path to the file to be attached
    """
    # Convert single email to list if necessary
    if isinstance(receiver_emails, str):
        receiver_emails = [receiver_emails]
    
    # Remove any empty strings and whitespace
    receiver_emails = [email.strip() for email in receiver_emails if email.strip()]
    
    if not receiver_emails:
        print("No valid receiver emails provided")
        return False
    
    try:
        # Create the email message
        email_message = MIMEMultipart()
        email_message['From'] = sender_email
        email_message['To'] = ', '.join(receiver_emails)  # Join all recipients with commas
        email_message['Subject'] = subject
        
        # Attach the message body
        email_message.attach(MIMEText(message, 'plain'))
        
        # Attach the file
        with open(attachment_path, 'rb') as file:
            part = MIMEApplication(file.read(), Name=os.path.basename(attachment_path))
            part['Content-Disposition'] = f'attachment; filename="{os.path.basename(attachment_path)}"'
            email_message.attach(part)
        
        # Send the email
        with smtplib.SMTP("smtp.gmail.com", 587) as server:
            server.starttls()
            server.login(sender_email, sender_password)
            server.send_message(email_message)
        
        print(f"Email sent successfully to: {', '.join(receiver_emails)}")
        return True
    
    except Exception as e:
        print(f"Error sending email: {e}")
        return False

def main():
    """
    Main function to process temperature data for multiple vehicles and generate report
    """
    end_date = datetime.now().replace(hour=16, minute=0, second=0, microsecond=0) - timedelta(days=1)
    start_date = (end_date - timedelta(days=1)).strftime('%Y-%m-%d %H:%M')
    end_date_str = end_date.strftime('%Y-%m-%d %H:%M')
    
    vehicles_data = []
    
    # Process each vehicle
    for device_id, plate_number in CONFIG['VEHICLES'].items():
        api_url = f"https://fms2.gpsbox.mn/api/api.php?api=user&key={CONFIG['GPS_API_KEY']}&cmd=OBJECT_GET_MESSAGES,{device_id},{start_date},{end_date_str},0.01"
        
        try:
            response = requests.get(api_url)
            json_data = response.json()
            
            # Parse and process data for this vehicle
            sensor_data = parse_gps_temp_door(json_data, plate_number)
            
            vehicle_data = {
                'plate_number': plate_number,
                'storage_temps': process_temperature_data(sensor_data['storage_temp']),
                'bag_temps': process_temperature_data(sensor_data['bag_temp']),
                'door_events': process_door_events(sensor_data['door'], 
                                                sensor_data['storage_temp'],
                                                sensor_data['bag_temp'])
            }
            
            vehicles_data.append(vehicle_data)
            
        except Exception as e:
            print(f"Error processing vehicle {plate_number} (ID: {device_id}): {e}")
    
    if vehicles_data:
        # Generate Excel report
        report_file = export_to_excel(vehicles_data)
        
        # Send email
        send_email_with_attachment(
            CONFIG['SENDER_EMAIL'],
            CONFIG['SENDER_PASSWORD'],
            CONFIG['RECEIVER_EMAILS'],
            f"Temperature Analysis Report - {end_date_str}",
            f"Attached is the temperature analysis report for all vehicles from {start_date} to {end_date_str}.",
            report_file
        )
    else:
        print("No data was processed successfully for any vehicle.")

if __name__ == "__main__":
    main()
