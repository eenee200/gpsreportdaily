import os
import sys
import json
import pandas as pd
import requests
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMENonMultipart
from email.mime.text import MIMEText
from email.mime.application import MIMEApplication
from datetime import datetime, timedelta
RECEIVER_EMAILS="dorjsuren.n@gmail.com,lhagvabayar.a@monospharmatrade.mn,undrakh.b@monospharmatrade.mn,anuerdene.b@monospharmatrade.mn,ayurzana.s@monospharmatrade.mn,munkhtamir.b@monospharmatrade.mn,baatarkhuu@monospharmatrade.mn,ariuntungalag@monospharmatrade.mn,narmandakh.b@monospharmatrade.mn"

# Configuration (Replace with your actual details)
CONFIG = {
    'GPS_API_KEY': os.environ.get('GPS_API_KEY'),
    'SENDER_EMAIL': os.environ.get('SENDER_EMAIL'),
    'SENDER_PASSWORD': os.environ.get('SENDER_PASSWORD'),
    'RECEIVER_EMAILS': RECEIVER_EMAILS.split(','),
    'REPORT_FREQUENCY_DAYS': 1,
    'VEHICLES': {
        '866069068751245': '5476УКК ОН',
        '867747074781160': '3049УАС МУБ',
        '866069068899358': '6461УНЯ ОН',
        '866069068945011': '7107УБГ ОН',
        '866069064383431': '7228УКР МУБ',
        '867747074781186': '7228УКС МУБ',
        '867747074781053': '7538УАМ',
        '868373075447468': '7204УЕУ'
    }
}

def parse_gps_temp_humidity(json_data, plate_number):
    """
    Parse GPS tracking data to extract refrigerator temperature and humidity.
    
    :param json_data: List of GPS tracking entries
    :param plate_number: Vehicle plate number
    :return: Dictionary with parsed sensor data
    """
    sensor_data = {
        'refrigerator_temp': [],
        'humidity': [],
        'plate_number': plate_number
    }
    last_valid_humidity = 0
    last_valid_temp = 0
    
    for entry in json_data:
        timestamp = datetime.strptime(entry[0], '%Y-%m-%d %H:%M:%S') + timedelta(hours=8)
        
        if 'io10800' in entry[6]:
            refrigerator_temp = float(entry[6]['io10800']) / 100
            if refrigerator_temp == 250:
                refrigerator_temp = last_valid_temp
            else:
                last_valid_temp = refrigerator_temp
            sensor_data['refrigerator_temp'].append({
                'timestamp': timestamp,
                'value': refrigerator_temp
            })
        
        if 'io10804' in entry[6]:
            humidity = float(entry[6]['io10804'])
            if humidity == 250:
                humidity = last_valid_humidity
            else:
                last_valid_humidity = humidity
            sensor_data['humidity'].append({
                'timestamp': timestamp,
                'value': humidity
            })
    
    return sensor_data

def fetch_vehicle_data(device_id, start_date, end_date, api_key):
    """
    Fetch and process data for a single vehicle.
    
    :param device_id: GPS device identifier
    :param start_date: Start date for data retrieval
    :param end_date: End date for data retrieval
    :param api_key: API authentication key
    :return: Processed sensor data for the vehicle
    """
    base_url = "https://fms2.gpsbox.mn/api/api.php"
    params = {
        "api": "user",
        "key": api_key,
        "cmd": f"OBJECT_GET_MESSAGES,{device_id},{start_date} 00:00:00,{end_date} 00:00:00,0.01"
    }
    
    api_url = f"{base_url}?api={params['api']}&key={params['key']}&cmd={params['cmd']}"
    
    try:
        response = requests.get(api_url)
        response.raise_for_status()
        json_data = response.json()
        return parse_gps_temp_humidity(json_data, CONFIG['VEHICLES'][device_id])
    except requests.RequestException as e:
        print(f"Error fetching GPS data for device {device_id}: {e}")
        return None

def process_sensor_data(sensor_data, interval_minutes=5):
    """
    Process sensor data into specified minute intervals.
    Returns empty list if no data is available.
    """
    # Check for empty or None input
    if not sensor_data:
        print("Warning: Empty sensor data received")
        return []
    
    try:
        # Create DataFrame
        df = pd.DataFrame(sensor_data)
        
        # Check if required columns exist
        if 'timestamp' not in df.columns or 'value' not in df.columns:
            print(f"Warning: Missing required columns. Available columns: {df.columns.tolist()}")
            return []
            
        # Convert timestamp to datetime and create intervals
        df['timestamp'] = pd.to_datetime(df['timestamp'])
        df['interval_timestamp'] = df['timestamp'].dt.floor(f'{interval_minutes}min')
        
        # Group by interval and calculate statistics
        grouped_data = df.groupby('interval_timestamp').agg({
            'value': ['mean', 'min', 'max', 'count']
        }).reset_index()
        
        # Flatten column names
        grouped_data.columns = ['timestamp', 'mean', 'min', 'max', 'count']
        
        # Round numeric values
        for col in ['mean', 'min', 'max']:
            grouped_data[col] = grouped_data[col].round(2)
        
        return grouped_data.to_dict('records')
        
    except Exception as e:
        print(f"Error processing sensor data: {e}")
        return []


def calculate_daily_averages(sensor_data):
    """Calculate daily averages for sensor data."""
    df = pd.DataFrame(sensor_data)
    df['date'] = pd.to_datetime(df['timestamp']).dt.date
    
    daily_averages = df.groupby('date').agg({
        'mean': ['mean', 'min', 'max']
    }).reset_index()
    
    daily_averages.columns = ['date', 'average', 'minimum', 'maximum']
    daily_averages['date'] = daily_averages['date'].astype(str)
    return daily_averages.to_dict('records')

def export_to_excel(vehicles_data, output_file='gps_sensor_analysis.xlsx'):
    """
    Export multi-vehicle sensor analysis to Excel.
    
    :param vehicles_data: Dictionary containing processed data for each vehicle
    :param output_file: Excel file path
    """
    import openpyxl
    from openpyxl.styles import Border, Side, Alignment
    
    workbook = openpyxl.Workbook()
    
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')
    
    # Create 5-minute data sheets for each vehicle
    for vehicle_id, data in vehicles_data.items():
        sheet_name = f"{data['plate_number']}_5min"
        if len(sheet_name) > 31:  # Excel sheet name length limitation
            sheet_name = sheet_name[:31]
        
        sheet = workbook.create_sheet(title=sheet_name)
        
        # Headers
        sheet.append(['Plate Number', 'Timestamp', 'Refrigerator Temp', 'Humidity'])
        
        # Check if data exists
        temp_data = data['refrigerator_temp']
        humidity_data = data['humidity']
        
        if not temp_data and not humidity_data:
            # If no data available, add a message
            sheet.append([data['plate_number'], 'No data available', '-', '-'])
        else:
            # Combine temperature and humidity data
            max_length = max(len(temp_data), len(humidity_data))
            
            for i in range(max_length):
                temp_value = round(temp_data[i]['mean'], 2) if i < len(temp_data) else '-'
                humidity_value = round(humidity_data[i]['mean'], 2) if i < len(humidity_data) else '-'
                timestamp = (temp_data[i]['timestamp'] if i < len(temp_data) 
                            else humidity_data[i]['timestamp'])
                
                sheet.append([
                    data['plate_number'] if i == 0 else '',
                    timestamp,
                    temp_value,
                    humidity_value
                ])
    
    # Create daily averages sheet
    daily_avg_sheet = workbook.create_sheet(title="Daily_Averages")
    daily_avg_sheet.append([
        'Plate Number',
        'Date',
        'Refrigerator Temp Avg', 'Refrigerator Temp Min', 'Refrigerator Temp Max',
        'Humidity Avg', 'Humidity Min', 'Humidity Max'
    ])
    
    # Add daily averages for each vehicle
    for vehicle_id, data in vehicles_data.items():
        temp_data = data['refrigerator_temp']
        humidity_data = data['humidity']
        
        if not temp_data and not humidity_data:
            # If no data available, add a message
            daily_avg_sheet.append([
                data['plate_number'],
                datetime.now().strftime('%Y-%m-%d'),
                'No data', 'No data', 'No data',
                'No data', 'No data', 'No data'
            ])
        else:
            temp_daily = calculate_daily_averages(temp_data) if temp_data else []
            humidity_daily = calculate_daily_averages(humidity_data) if humidity_data else []
            
            if not temp_daily and not humidity_daily:
                daily_avg_sheet.append([
                    data['plate_number'],
                    datetime.now().strftime('%Y-%m-%d'),
                    'No data', 'No data', 'No data',
                    'No data', 'No data', 'No data'
                ])
            else:
                for i in range(len(temp_daily)):
                    daily_avg_sheet.append([
                        data['plate_number'],
                        temp_daily[i]['date'],
                        round(temp_daily[i]['average'], 2),
                        round(temp_daily[i]['minimum'], 2),
                        round(temp_daily[i]['maximum'], 2),
                        round(humidity_daily[i]['average'], 2) if i < len(humidity_daily) else 'No data',
                        round(humidity_daily[i]['minimum'], 2) if i < len(humidity_daily) else 'No data',
                        round(humidity_daily[i]['maximum'], 2) if i < len(humidity_daily) else 'No data'
                    ])
    
    # Remove default sheet if it exists
    if 'Sheet' in workbook.sheetnames:
        workbook.remove(workbook['Sheet'])
    
    # Apply formatting to all sheets
    for sheet in workbook.sheetnames:
        ws = workbook[sheet]
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for cell in row:
                cell.border = border
                cell.alignment = alignment
        
        # Adjust column widths
        for col in ws.columns:
            max_length = 0
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[col[0].column_letter].width = adjusted_width
    
    workbook.save(output_file)
    return output_file



def send_email_with_attachment(sender_email, sender_password, receiver_emails, 
                             subject, message, attachment_path):
    """
    Send an email with an Excel file attachment to multiple recipients.
    
    :param sender_email: Sender's email address
    :param sender_password: Sender's email password
    :param receiver_emails: List of recipient email addresses
    :param subject: Email subject
    :param message: Email body text
    :param attachment_path: Path to the Excel file to attach
    """
    try:
        # Create email message
        email_message = MIMEMultipart()
        email_message['From'] = sender_email
        email_message['To'] = ', '.join(receiver_emails)  # Join all recipients with commas
        email_message['Subject'] = subject

        # Attach message body
        email_message.attach(MIMEText(message, 'plain'))

        # Attach Excel file
        with open(attachment_path, 'rb') as file:
            part = MIMEApplication(file.read(), Name=os.path.basename(attachment_path))
            part['Content-Disposition'] = f'attachment; filename="{os.path.basename(attachment_path)}"'
            email_message.attach(part)

        # Send email
        with smtplib.SMTP("smtp.gmail.com", 587) as server:
            server.starttls()
            server.login(sender_email, sender_password)
            server.send_message(email_message)
        
        print(f"Email sent successfully to {', '.join(receiver_emails)}")
        return True
    except Exception as e:
        print(f"Error sending email: {e}")
        return False

def main():
    """Main function to process multiple vehicles' data."""
    try:
        end_date = datetime.now().replace(hour=16, minute=0, second=0, microsecond=0) - timedelta(days=1)
        start_date = (end_date - timedelta(days=1)).strftime('%Y-%m-%d %H:%M')
        end_date = end_date.strftime('%Y-%m-%d %H:%M')
        
        vehicles_data = {}
        
        # Fetch and process data for each vehicle
        for device_id in CONFIG['VEHICLES'].keys():
            try:
                print(f"Processing device {device_id} ({CONFIG['VEHICLES'][device_id]})")
                
                sensor_data = fetch_vehicle_data(
                    device_id,
                    start_date,
                    end_date,
                    CONFIG['GPS_API_KEY']
                )
                
                # Initialize data structure even if no sensor data
                processed_data = {
                    'plate_number': CONFIG['VEHICLES'][device_id],
                    'refrigerator_temp': [],
                    'humidity': []
                }
                
                # Process sensor data if available
                if sensor_data and isinstance(sensor_data, dict):
                    temp_data = process_sensor_data(sensor_data.get('refrigerator_temp', []))
                    humidity_data = process_sensor_data(sensor_data.get('humidity', []))
                    
                    if temp_data or humidity_data:
                        processed_data['refrigerator_temp'] = temp_data
                        processed_data['humidity'] = humidity_data
                
                vehicles_data[device_id] = processed_data
                
            except Exception as e:
                print(f"Error processing device {device_id}: {e}")
                # Add empty data structure for failed device
                vehicles_data[device_id] = {
                    'plate_number': CONFIG['VEHICLES'][device_id],
                    'refrigerator_temp': [],
                    'humidity': []
                }
                continue
        
        if not vehicles_data:
            print("No data retrieved for any vehicle.")
            return
        
        # Generate Excel report
        report_file = export_to_excel(vehicles_data)
        
        # Send email with report to all recipients
        send_email_with_attachment(
            CONFIG['SENDER_EMAIL'],
            CONFIG['SENDER_PASSWORD'],
            CONFIG['RECEIVER_EMAILS'],  # Now passing the list of all recipients
            f"Multi-Vehicle GPS Sensor Report - {end_date}",
            f"Attached is the GPS sensor report for all vehicles from {start_date} to {end_date}.",
            report_file
        )
    
    except Exception as e:
        print(f"Error in main function: {e}")

if __name__ == "__main__":
    main()
